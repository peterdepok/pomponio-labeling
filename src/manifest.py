"""Manifest spreadsheet generation for animal closeout.

Generates an Excel file per animal with:
    Columns: SKU, Product Name, Quantity, Individual Weights, Total Weight
    One row per unique product.
    Summary totals at bottom.
"""

import logging
import os
from datetime import datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.database import Database

logger = logging.getLogger(__name__)

# Output directory for manifests
MANIFEST_DIR = "data/manifests"

# Styles
HEADER_FONT = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
BODY_FONT = Font(name="Calibri", size=11)
TOTAL_FONT = Font(name="Calibri", size=12, bold=True)
TOTAL_FILL = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def generate_manifest(
    db: Database,
    animal_id: int,
    output_dir: str = MANIFEST_DIR,
) -> Optional[str]:
    """Generate a manifest spreadsheet for an animal.

    Args:
        db: Database instance.
        animal_id: ID of the animal to generate manifest for.
        output_dir: Directory to save the spreadsheet.

    Returns:
        Path to the generated file, or None on failure.
    """
    animal = db.get_animal(animal_id)
    if animal is None:
        logger.error("Animal %d not found", animal_id)
        return None

    data = db.get_animal_manifest_data(animal_id)
    if not data:
        logger.warning("No packages for animal %d", animal_id)
        return None

    # Ensure output directory exists
    os.makedirs(output_dir, exist_ok=True)

    # Build filename
    safe_name = animal["name"].replace("/", "-").replace("\\", "-").replace(" ", "_")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"manifest_{safe_name}_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)

    wb = Workbook()
    ws = wb.active
    ws.title = "Manifest"

    # Title rows
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value = f"Pomponio Ranch Manifest: {animal['name']}"
    title_cell.font = Font(name="Calibri", size=16, bold=True)
    title_cell.alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:E2")
    date_cell = ws["A2"]
    date_cell.value = f"Generated: {datetime.now().strftime('%m/%d/%Y %I:%M %p')}"
    date_cell.font = Font(name="Calibri", size=10, italic=True)
    date_cell.alignment = Alignment(horizontal="center")

    # Headers
    headers = ["SKU", "Product Name", "Qty", "Individual Weights (lb)", "Total Weight (lb)"]
    header_row = 4
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows
    row = header_row + 1
    grand_qty = 0
    grand_weight = 0.0

    for item in data:
        weights_str = ", ".join(f"{w:.2f}" for w in item["weights"])

        ws.cell(row=row, column=1, value=item["sku"]).font = BODY_FONT
        ws.cell(row=row, column=2, value=item["product_name"]).font = BODY_FONT
        ws.cell(row=row, column=3, value=item["quantity"]).font = BODY_FONT
        ws.cell(row=row, column=4, value=weights_str).font = BODY_FONT
        ws.cell(row=row, column=5, value=round(item["total_weight"], 2)).font = BODY_FONT

        for col in range(1, 6):
            ws.cell(row=row, column=col).border = THIN_BORDER

        ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=5).alignment = Alignment(horizontal="right")
        ws.cell(row=row, column=5).number_format = "0.00"

        grand_qty += item["quantity"]
        grand_weight += item["total_weight"]
        row += 1

    # Totals row
    ws.cell(row=row, column=1, value="TOTAL").font = TOTAL_FONT
    ws.cell(row=row, column=1).fill = TOTAL_FILL
    ws.cell(row=row, column=2).fill = TOTAL_FILL
    ws.cell(row=row, column=3, value=grand_qty).font = TOTAL_FONT
    ws.cell(row=row, column=3).fill = TOTAL_FILL
    ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=4).fill = TOTAL_FILL
    ws.cell(row=row, column=5, value=round(grand_weight, 2)).font = TOTAL_FONT
    ws.cell(row=row, column=5).fill = TOTAL_FILL
    ws.cell(row=row, column=5).alignment = Alignment(horizontal="right")
    ws.cell(row=row, column=5).number_format = "0.00"

    for col in range(1, 6):
        ws.cell(row=row, column=col).border = THIN_BORDER

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 50
    ws.column_dimensions["E"].width = 16

    # Freeze header row
    ws.freeze_panes = "A5"

    wb.save(filepath)
    logger.info("Manifest saved: %s (%d SKUs, %d packages)", filepath, len(data), grand_qty)
    return filepath
