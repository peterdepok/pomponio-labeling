"""Tests for manifest spreadsheet generation."""

import os
import tempfile
import pytest

from openpyxl import load_workbook

from src.database import Database
from src.manifest import generate_manifest


@pytest.fixture
def db():
    d = Database(":memory:")
    d.connect()
    csv_path = os.path.join(os.path.dirname(__file__), "..", "data", "pomponio_skus.csv")
    d.import_products_from_csv(csv_path)
    return d


@pytest.fixture
def populated_db(db):
    """Database with an animal, box, and packages."""
    aid = db.create_animal("Beef #1 - Test")
    bid = db.create_box(aid)

    p1 = db.get_product_by_sku("00100")
    p2 = db.get_product_by_sku("00123")

    db.create_package(p1["id"], aid, bid, 1.52, "000100001525")
    db.create_package(p1["id"], aid, bid, 2.00, "000100002008")
    db.create_package(p2["id"], aid, bid, 1.00, "001230010006")

    return db, aid


class TestManifestGeneration:

    def test_generates_file(self, populated_db):
        db, aid = populated_db
        with tempfile.TemporaryDirectory() as tmpdir:
            path = generate_manifest(db, aid, output_dir=tmpdir)
            assert path is not None
            assert os.path.exists(path)
            assert path.endswith(".xlsx")

    def test_file_content(self, populated_db):
        db, aid = populated_db
        with tempfile.TemporaryDirectory() as tmpdir:
            path = generate_manifest(db, aid, output_dir=tmpdir)
            wb = load_workbook(path)
            ws = wb.active

            # Check title
            assert "Beef #1" in ws["A1"].value

            # Check headers in row 4
            assert ws.cell(row=4, column=1).value == "SKU"
            assert ws.cell(row=4, column=2).value == "Product Name"
            assert ws.cell(row=4, column=3).value == "Qty"

            # Check data rows (2 SKUs)
            skus = set()
            for row in range(5, 7):
                sku = ws.cell(row=row, column=1).value
                if sku and sku != "TOTAL":
                    skus.add(sku)
            assert "00100" in skus
            assert "00123" in skus

    def test_totals_row(self, populated_db):
        db, aid = populated_db
        with tempfile.TemporaryDirectory() as tmpdir:
            path = generate_manifest(db, aid, output_dir=tmpdir)
            wb = load_workbook(path)
            ws = wb.active

            # Find totals row
            total_row = None
            for row in range(5, 20):
                if ws.cell(row=row, column=1).value == "TOTAL":
                    total_row = row
                    break

            assert total_row is not None
            assert ws.cell(row=total_row, column=3).value == 3  # 3 packages
            assert abs(ws.cell(row=total_row, column=5).value - 4.52) < 0.01

    def test_nonexistent_animal(self, db):
        with tempfile.TemporaryDirectory() as tmpdir:
            result = generate_manifest(db, 9999, output_dir=tmpdir)
            assert result is None

    def test_empty_animal(self, db):
        aid = db.create_animal("Empty Animal")
        with tempfile.TemporaryDirectory() as tmpdir:
            result = generate_manifest(db, aid, output_dir=tmpdir)
            assert result is None
